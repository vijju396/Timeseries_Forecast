import csv
import hashlib
import os
import warnings
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path


RAW_SCHEMA_VERSION = "raw-schema-v1"
DEFAULT_PREVIEW_LIMIT = 10
MAX_PREVIEW_LIMIT = 50
SCHEMA_SAMPLE_ROWS = 100


class RawPreviewError(ValueError):
    pass


def source_file_hash(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def raw_artifact_id(dataset_id, source_hash):
    value = f"{RAW_SCHEMA_VERSION}|{dataset_id}|{source_hash}"
    return hashlib.sha256(value.encode("utf-8")).hexdigest()[:24]


def build_raw_preview(path, original_filename, dataset_id, limit=DEFAULT_PREVIEW_LIMIT, source_hash=None):
    limit = normalize_limit(limit)
    source_hash = source_hash or source_file_hash(path)
    frame, source_headers, parse_metadata, total_rows, row_count_exact = read_dataframe(
        path, nrows=max(limit, SCHEMA_SAMPLE_ROWS), include_count=True
    )
    schema, warnings = build_schema(source_headers, frame)
    if parse_metadata.get("encoding_warning"):
        warnings.append(parse_metadata["encoding_warning"])
    artifact_id = raw_artifact_id(dataset_id, source_hash)
    preview = {
        "ok": True,
        "dataset_id": dataset_id,
        "raw_artifact_id": artifact_id,
        "source_file_hash": source_hash,
        "original_filename": original_filename,
        "preview_type": "raw",
        "schema_version": RAW_SCHEMA_VERSION,
        "columns": [public_column(column) for column in schema],
        "rows": frame_to_rows(frame.head(limit), schema),
        "preview_row_count": min(limit, len(frame)),
        "schema_sample_row_count": len(frame),
        "total_raw_rows": total_rows,
        "row_count_exact": row_count_exact,
        "parse_metadata": parse_metadata,
        "warnings": warnings,
    }
    preview["preview_cache_key"] = hashlib.sha256(
        f"{dataset_id}|{artifact_id}|{source_hash}|raw|{limit}|{RAW_SCHEMA_VERSION}".encode("utf-8")
    ).hexdigest()
    return preview, schema, frame


def reload_raw_preview(path, dataset, limit=DEFAULT_PREVIEW_LIMIT):
    limit = normalize_limit(limit)
    stored_hash = dataset.get("source_file_hash")
    if not stored_hash or dataset.get("raw_artifact_id") != raw_artifact_id(dataset.get("id"), stored_hash):
        raise RawPreviewError("Raw preview ownership metadata is invalid. Upload the dataset again.")
    source = Path(path)
    expected_size = dataset.get("raw_file_size")
    expected_mtime = dataset.get("raw_file_mtime_ns")
    stat = source.stat()
    if expected_size is not None and int(expected_size) != stat.st_size:
        raise RawPreviewError("The immutable raw upload has changed. Upload the dataset again.")
    if expected_mtime is not None and int(expected_mtime) != stat.st_mtime_ns:
        if source_file_hash(source) != stored_hash:
            raise RawPreviewError("The immutable raw upload has changed. Upload the dataset again.")
    preview, schema, _frame = build_raw_preview(
        source,
        dataset.get("original_filename") or source.name,
        dataset.get("id"),
        limit=limit,
        source_hash=stored_hash,
    )
    if preview["raw_artifact_id"] != dataset.get("raw_artifact_id"):
        raise RawPreviewError("Raw preview does not belong to the active dataset.")
    return preview, schema


def read_dataframe(path, nrows=None, include_count=False):
    import pandas as pd

    source = Path(path)
    if not source.exists() or not source.is_file():
        raise RawPreviewError("Raw upload artifact is unavailable.")
    if source.suffix.lower() == ".xlsx":
        try:
            raw = pd.read_excel(source, header=None, nrows=None if nrows is None else nrows + 1)
        except Exception as exc:
            raise RawPreviewError(f"Excel parsing failed: {exc}") from exc
        if raw.empty:
            raise RawPreviewError("The uploaded workbook is empty.")
        source_headers = [_header_text(value, index) for index, value in enumerate(raw.iloc[0].tolist())]
        internal_headers = dedupe_source_names(source_headers)
        frame = raw.iloc[1:].reset_index(drop=True)
        frame.columns = internal_headers
        total = max(0, len(raw) - 1) if nrows is None else None
        exact = nrows is None
        if include_count and nrows is not None:
            try:
                from openpyxl import load_workbook

                workbook = load_workbook(source, read_only=True, data_only=True)
                try:
                    total = max(0, workbook.active.max_row - 1)
                    exact = True
                finally:
                    workbook.close()
            except Exception:
                # Preview and schema remain usable even when an exact workbook row count is unavailable.
                total, exact = None, False
        return frame, source_headers, {"format": "xlsx", "encoding": None, "delimiter": None, "quote_character": None}, total, exact

    spec = detect_text_format(source)
    if nrows is None:
        try:
            source_headers, _empty = _sample_text_rows(source, spec, 0)
            internal_headers = dedupe_source_names(source_headers)
            separator = spec.get("delimiter") or "\x1f"
            frame = pd.read_csv(
                source,
                sep=separator,
                quotechar=spec.get("quote_character") or '"',
                encoding=spec["encoding"],
                header=0,
                names=internal_headers,
            )
        except Exception as exc:
            raise RawPreviewError(f"Delimited-text parsing failed: {exc}") from exc
        return frame, source_headers, spec, len(frame) if include_count else None, bool(include_count)
    try:
        source_headers, sample_rows = _sample_text_rows(source, spec, None if nrows is None else nrows)
    except (csv.Error, UnicodeError, OSError) as exc:
        raise RawPreviewError(f"Delimited-text parsing failed: {exc}") from exc
    internal_headers = dedupe_source_names(source_headers)
    frame = pd.DataFrame(sample_rows, columns=internal_headers)
    total = _count_text_rows(source, spec) if include_count else None
    exact = total is not None
    return frame, source_headers, spec, total, exact


def detect_text_format(path):
    source = Path(path)
    raw = source.read_bytes()[:131072]
    if not raw:
        raise RawPreviewError("The uploaded file is empty.")
    if b"\x00" in raw and not (raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff")):
        raise RawPreviewError("The uploaded text file contains binary data.")
    encoding, warning = _detect_encoding(raw)
    try:
        sample = raw.decode(encoding)
    except UnicodeDecodeError as exc:
        raise RawPreviewError("The uploaded text encoding could not be decoded safely.") from exc
    sample = sample.lstrip("\ufeff")
    delimiter = None
    quote_character = '"'
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delimiter = dialect.delimiter
        quote_character = dialect.quotechar or '"'
    except csv.Error:
        if any(character in sample for character in (",", "\t", ";", "|")):
            raise RawPreviewError("Delimiter detection was ambiguous. Use a consistently delimited CSV, TSV, or text file.")
    return {
        "format": "delimited_text",
        "encoding": encoding,
        "delimiter": delimiter,
        "quote_character": quote_character,
        "encoding_warning": warning,
    }


def dedupe_source_names(source_headers):
    seen = Counter()
    internal = []
    for index, source_name in enumerate(source_headers):
        base = str(source_name) if source_name != "" else f"column_{index + 1}"
        seen[base] += 1
        internal.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return internal


def build_schema(source_headers, frame):
    internal_headers = dedupe_source_names(source_headers)
    duplicate_names = {name for name, count in Counter(source_headers).items() if count > 1}
    warnings = []
    if duplicate_names:
        warnings.append("Duplicate source headers were preserved with stable column IDs for mapping.")
    if any(name == "" for name in source_headers):
        warnings.append("One or more source headers are empty; column positions are preserved.")
    schema = []
    for position, (source_name, internal_name) in enumerate(zip(source_headers, internal_headers)):
        values = frame[internal_name].tolist() if internal_name in frame else []
        physical_type, semantic, confidence = infer_column(values)
        examples = []
        for value in values:
            safe = json_value(value)
            if safe is not None and safe not in examples:
                examples.append(safe)
            if len(examples) == 3:
                break
        null_count = sum(json_value(value) is None for value in values)
        schema.append({
            "column_id": f"col_{position}",
            "source_name": source_name,
            "display_name": source_name,
            "internal_name": internal_name,
            "position": position,
            "physical_type": physical_type,
            "semantic_suggestion": semantic,
            "nullable": null_count > 0,
            "null_count": null_count,
            "example_values": examples,
            "parsing_confidence": confidence,
            "duplicate_source_name": source_name in duplicate_names,
        })
    return schema, warnings


def infer_column(values):
    import pandas as pd

    clean = [value for value in values if json_value(value) is not None]
    if not clean:
        return "unknown", "unknown", 0.0
    text = [str(value).strip() for value in clean]
    boolean_ratio = sum(value.lower() in {"true", "false", "yes", "no"} for value in text) / len(text)
    numeric = pd.to_numeric(pd.Series(text), errors="coerce")
    numeric_ratio = float(numeric.notna().mean())
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        dates = pd.to_datetime(pd.Series(text), errors="coerce")
    date_ratio = float(dates.notna().mean())
    if boolean_ratio >= 0.9:
        return "boolean", "boolean", round(boolean_ratio, 3)
    if numeric_ratio >= 0.9:
        return "number", "numeric", round(numeric_ratio, 3)
    if date_ratio >= 0.9:
        return "datetime", "timestamp", round(date_ratio, 3)
    distinct_ratio = len(set(text)) / max(len(text), 1)
    semantic = "category" if distinct_ratio <= 0.5 else "identifier_or_text"
    return "string", semantic, round(max(0.5, 1 - distinct_ratio / 2), 3)


def frame_to_rows(frame, schema):
    rows = []
    for _index, source_row in frame.iterrows():
        rows.append({column["column_id"]: json_value(source_row.get(column["internal_name"])) for column in schema})
    return rows


def public_column(column):
    return {key: value for key, value in column.items() if key != "internal_name"}


def normalize_limit(value):
    try:
        return max(1, min(int(value), MAX_PREVIEW_LIMIT))
    except (TypeError, ValueError):
        return DEFAULT_PREVIEW_LIMIT


def upload_timestamp():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def json_value(value):
    if value is None:
        return None
    try:
        import pandas as pd

        missing = pd.isna(value)
        if isinstance(missing, bool) and missing:
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(value, "item"):
        value = value.item()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _sample_text_rows(path, spec, nrows):
    encoding = spec["encoding"]
    delimiter = spec.get("delimiter")
    quote_character = spec.get("quote_character") or '"'
    with Path(path).open("r", encoding=encoding, newline="") as handle:
        if delimiter:
            reader = csv.reader(handle, delimiter=delimiter, quotechar=quote_character)
        else:
            reader = ([line.rstrip("\r\n")] for line in handle)
        try:
            header = next(reader)
        except StopIteration as exc:
            raise RawPreviewError("The uploaded file is empty.") from exc
        header = [str(value) for value in header]
        if header:
            header[0] = header[0].lstrip("\ufeff")
        if not header or len(header) > 1000:
            raise RawPreviewError("The uploaded header is empty or exceeds the safe column limit.")
        rows = []
        max_rows = nrows
        if max_rows == 0:
            return header, rows
        for row_number, row in enumerate(reader, start=2):
            if len(row) != len(header):
                raise RawPreviewError(f"Row {row_number} has {len(row)} fields; expected {len(header)} from the header.")
            rows.append([None if value == "" else value for value in row])
            if max_rows is not None and len(rows) >= max_rows:
                break
    return header, rows


def _count_text_rows(path, spec):
    max_bytes = int(os.getenv("RAW_PREVIEW_EXACT_COUNT_MAX_BYTES", str(10 * 1024 * 1024)))
    source = Path(path)
    if source.stat().st_size > max_bytes:
        return None
    encoding = spec["encoding"]
    delimiter = spec.get("delimiter")
    with source.open("r", encoding=encoding, newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter, quotechar=spec.get("quote_character") or '"') if delimiter else handle
        return max(0, sum(1 for _row in reader) - 1)


def _detect_encoding(raw):
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig", None
    if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
        return "utf-16", None
    try:
        raw.decode("utf-8")
        return "utf-8", None
    except UnicodeDecodeError:
        return "cp1252", "The source encoding was normalized from Windows-1252."


def _header_text(value, _position):
    safe = json_value(value)
    return "" if safe is None else str(safe).lstrip("\ufeff")
